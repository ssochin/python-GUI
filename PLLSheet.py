import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import math


class PLLSheet:

    def __init__(self):
        self.filename = "empty"
        self.wb = "empty"
        self.sheet = "empty"
        self.framerate = 600

    # creates spreadsheet that stores entire calibration table
    def generate_caltable_spreadsheet(self, filename, freqrange, caltable):
        # create spreadsheet
        wb = Workbook()  # create spreadsheet
        wb.save(filename)  # save workbook
        sheet = wb.active  # grab active sheet

        # create titles in caltable spreadsheet
        sheet['A1'] = "Frequency"
        sheet['B1'] = "R46"
        sheet['C1'] = "R45"
        sheet['D1'] = "R44"
        sheet['E1'] = "R37"
        sheet['F1'] = "R36"
        sheet['G1'] = "R20"
        sheet['H1'] = "R19"
        sheet['I1'] = "R14"
        sheet['J1'] = "R8"
        sheet['K1'] = "R0"



        # fill in the spreadsheet
        row = 2
        for j in freqrange:
            sheet["A" + str(row)] = str(j)  # frequency
            sheet["B" + str(row)] = hex(caltable[j][46])  # R46
            sheet["C" + str(row)] = hex(caltable[j][45])  # R45
            sheet["D" + str(row)] = hex(caltable[j][44])  # R44
            sheet["E" + str(row)] = hex(caltable[j][37])  # R37
            sheet["F" + str(row)] = hex(caltable[j][36])  # R36
            sheet["G" + str(row)] = hex(caltable[j][20])  # R20
            sheet["H" + str(row)] = hex(caltable[j][19])  # R19
            sheet["I" + str(row)] = hex(caltable[j][14])  # R14
            sheet["J" + str(row)] = hex(caltable[j][8])  # R8
            sheet["K" + str(row)] = hex(caltable[j][0])  # R0
            row = row + 1


        wb.save(filename)  # save file after modification


    #  create the spreadsheet and add skeleton
    def create_spreadsheet(self, sheet_name):
        self.filename = sheet_name  # save filename to object for later

        # open spreadsheet saved from GUI
        self.wb = Workbook()  # create spreadsheet
        self.wb.save(self.filename)  # save workbook
        self.sheet = self.wb.active  # grab active sheet

        # create basic skeleton in spreadsheet
        # titles
        self.sheet['A1'] = "mosiData"
        self.sheet['B1'] = "GPIO0"
        self.sheet['C1'] = "GPIO1"
        self.sheet['D1'] = "GPIO2"
        self.sheet['E1'] = "GPIO3"
        self.sheet['F1'] = "misoData"
        self.sheet['I1'] = "Timings"
        #timings column titles and values
        self.sheet['I2'] = "Frame Rate(ns)"
        self.sheet['J2'] = self.framerate
        self.sheet['I3'] = "N bits"
        self.sheet['J3'] = 24
        self.sheet['I4'] = "t_high(ns)"
        self.sheet['J4'] = 20
        self.sheet['I5'] = "t_low(ns)"
        self.sheet['J5'] = 20
        self.sheet['I6'] = "t_setup(ns)"
        self.sheet['J6'] = 20
        self.sheet['I7'] = "t_hold(ns)"
        self.sheet['J7'] = 20
        self.sheet['I8'] = "t_disable(ns)"
        self.sheet['J8'] = 100
        self.sheet['I9'] = "t_nextsampling_edge(ns)"
        self.sheet['J9'] = "0.0"
        self.sheet['I10'] = "t_enable(ns)"
        self.sheet['J10'] = 100
        self.sheet['I11'] = "t_min. synchigh time (ns)"
        self.sheet['J11'] = 20
        self.sheet['I12'] = "Polarity"
        self.sheet['J12'] = 0
        self.sheet['I13'] = "Phase"
        self.sheet['J13'] = 0
        self.sheet['I14'] = "ti0"
        self.sheet['J14'] = "0.0"
        self.sheet['I15'] = "tp0"
        self.sheet['J15'] = "0.0"
        self.sheet['I16'] = "ti1"
        self.sheet['J16'] = "0.0"
        self.sheet['I17'] = "tp1"
        self.sheet['J17'] = "0.0"
        self.sheet['I18'] = "ti2"
        self.sheet['J18'] = "0.0"
        self.sheet['I19'] = "tp2"
        self.sheet['J19'] = "0.0"
        self.sheet['I20'] = "ti3"
        self.sheet['J20'] = "0.0"
        self.sheet['I21'] = "tp3"
        self.sheet['J21'] = "0.0"
        self.sheet['I22'] = "Delay GPIO signals N frames"
        self.sheet['J22'] = 0
        self.sheet['I23'] = "Idle Value GPIO0"
        self.sheet['J23'] = 0
        self.sheet['I24'] = "Idle Value GPIO1"
        self.sheet['J24'] = 0
        self.sheet['I25'] = "Idle Value GPIO2"
        self.sheet['J25'] = 0
        self.sheet['I26'] = "Idle Value GPIO3"
        self.sheet['J26'] = 0
        self.sheet['I27'] = "Reading time"
        self.sheet['J27'] = 30
        self.sheet['I28'] = "Protocol"
        self.sheet['J28'] = "SPI"
        self.sheet['I29'] = "Resolution"
        self.sheet['J29'] = "5.0ns"

        self.wb.save(self.filename)  # save file after modification

    # add freq1 mosi data, delay and freq2 mosi data to spreadsheet
    def add_mosidata(self, delay, freq1, freq2, freq1_table, freq2_table):

        delayAtoB = math.ceil((delay*1000) / self.framerate)  # convert delay to ns then divide by frame rate and roundup
        delayBtoA = delayAtoB  # same delay

        # create data structure for freq 1 data
        R46 = hex((46 << 16) | freq1_table[46])  # add header to data from lookup table
        R46 = str(R46[2:])  # get rid of leading 0x
        R45 = hex((45 << 16) | freq1_table[45])  # add header to data from lookup table
        R45 = str(R45[2:])  # get rid of leading 0x
        R44 = hex((44 << 16) | freq1_table[44])  # add header to data from lookup table
        R44 = str(R44[2:])  # get rid of leading 0x
        R37 = hex((37 << 16) | freq1_table[37])  # add header to data from lookup table
        R37 = str(R37[2:])  # get rid of leading 0x
        R36 = hex((36 << 16) | freq1_table[36])  # add header to data from lookup table
        R36 = str(R36[2:])  # get rid of leading 0x
        R20 = hex((20 << 16) | freq1_table[20])  # add header to data from lookup table
        R20 = str(R20[2:])  # get rid of leading 0x
        R19 = hex((19 << 16) | freq1_table[19])  # add header to data from lookup table
        R19 = str(R19[2:])  # get rid of leading 0x
        R14 = hex((14 << 16) | freq1_table[14])  # add header to data from lookup table
        R14 = str(R14[2:])  # get rid of leading 0x
        R14 = '0' + R14  # pad with leading 0
        R8 = hex((8 << 16) | freq1_table[8])  # add header to data from lookup table
        R8 = str(R8[2:])  # get rid of leading 0x
        R8 = '0' + R8  # pad with leading 0
        R0 = hex((0 << 16) | freq1_table[0])  # add header to data from lookup table
        R0 = str(R0[2:])  # get rid of leading 0x
        R0 = '00' + R0  # pad with 2 leading 0's

        f1_data = [R46, R45, R44, R37, R36, R20, R19, R14, R8, R0]  # put freq1 data in array


        index = 2  # start at index 2

        # put data in spreadsheet
        for d in f1_data:
            self.sheet['A'+str(index)] = d
            self.sheet['B'+str(index)] = 0
            self.sheet['C'+str(index)] = 0
            self.sheet['D'+str(index)] = 0
            self.sheet['E'+str(index)] = 0
            index = index + 1

        # delay between A and B
        index = index + delayAtoB


        # create data structure for freq 2 data
        R46 = hex((46 << 16) | freq2_table[46])  # add header to data from lookup table
        R46 = str(R46[2:])  # get rid of leading 0x
        R45 = hex((45 << 16) | freq2_table[45])  # add header to data from lookup table
        R45 = str(R45[2:])  # get rid of leading 0x
        R44 = hex((44 << 16) | freq2_table[44])  # add header to data from lookup table
        R44 = str(R44[2:])  # get rid of leading 0x
        R37 = hex((37 << 16) | freq2_table[37])  # add header to data from lookup table
        R37 = str(R37[2:])  # get rid of leading 0x
        R36 = hex((36 << 16) | freq2_table[36])  # add header to data from lookup table
        R36 = str(R36[2:])  # get rid of leading 0x
        R20 = hex((20 << 16) | freq2_table[20])  # add header to data from lookup table
        R20 = str(R20[2:])  # get rid of leading 0x
        R19 = hex((19 << 16) | freq2_table[19])  # add header to data from lookup table
        R19 = str(R19[2:])  # get rid of leading 0x
        R14 = hex((14 << 16) | freq2_table[14])  # add header to data from lookup table
        R14 = str(R14[2:])  # get rid of leading 0x
        R14 = '0' + R14  # pad with leading 0
        R8 = hex((8 << 16) | freq2_table[8])  # add header to data from lookup table
        R8 = str(R8[2:])  # get rid of leading 0x
        R8 = '0' + R8  # pad with leading 0
        R0 = hex((0 << 16) | freq2_table[0])  # add header to data from lookup table
        R0 = str(R0[2:])  # get rid of leading 0x
        R0 = '00' + R0  # pad with 2 leading 0's

        f2_data = [R46, R45, R44, R37, R36, R20, R19, R14, R8, R0]  # put freq1 data in array

        # put data in spreadsheet
        for d in f2_data:
            self.sheet['A'+str(index)] = d
            self.sheet['B'+str(index)] = 0
            self.sheet['C'+str(index)] = 0
            self.sheet['D'+str(index)] = 0
            self.sheet['E'+str(index)] = 0
            index = index + 1

        # delay between B and A
        index = index + delayBtoA

        # frequency 1 again
        # put data in spreadsheet again
        for d in f1_data:
            self.sheet['A'+str(index)] = d
            self.sheet['B'+str(index)] = 0
            self.sheet['C'+str(index)] = 0
            self.sheet['D'+str(index)] = 0
            self.sheet['E'+str(index)] = 0
            index = index + 1

        # add labels to describe excel sheet
        self.sheet['M1'] = "Script by SH"

        self.sheet['M3'] = "Frequency 1 (MHz)"
        self.sheet['N3'] = freq1
        self.sheet['M4'] = "Frequency 2 (MHz)"
        self.sheet['N4'] = freq2
        self.sheet['M5'] = "Actual Delay (uS)"
        self.sheet['N5'] = delayBtoA*self.framerate/1000




        self.wb.save(self.filename)  # save file after modification



    # add freq1 mosi data, delay and freq2 mosi data to spreadsheet per frequency
    def add_mosidataPerFreq(self, freq1, freq1_table):

        # create data structure for freq 1 data
        R46 = hex((46 << 16) | freq1_table[46])  # add header to data from lookup table
        R46 = str(R46[2:])  # get rid of leading 0x
        R45 = hex((45 << 16) | freq1_table[45])  # add header to data from lookup table
        R45 = str(R45[2:])  # get rid of leading 0x
        R44 = hex((44 << 16) | freq1_table[44])  # add header to data from lookup table
        R44 = str(R44[2:])  # get rid of leading 0x
        R37 = hex((37 << 16) | freq1_table[37])  # add header to data from lookup table
        R37 = str(R37[2:])  # get rid of leading 0x
        R36 = hex((36 << 16) | freq1_table[36])  # add header to data from lookup table
        R36 = str(R36[2:])  # get rid of leading 0x
        R20 = hex((20 << 16) | freq1_table[20])  # add header to data from lookup table
        R20 = str(R20[2:])  # get rid of leading 0x
        R19 = hex((19 << 16) | freq1_table[19])  # add header to data from lookup table
        R19 = str(R19[2:])  # get rid of leading 0x
        R14 = hex((14 << 16) | freq1_table[14])  # add header to data from lookup table
        R14 = str(R14[2:])  # get rid of leading 0x
        R14 = '0' + R14  # pad with leading 0
        R8 = hex((8 << 16) | freq1_table[8])  # add header to data from lookup table
        R8 = str(R8[2:])  # get rid of leading 0x
        R8 = '0' + R8  # pad with leading 0
        R0 = hex((0 << 16) | freq1_table[0])  # add header to data from lookup table
        R0 = str(R0[2:])  # get rid of leading 0x
        R0 = '00' + R0  # pad with 2 leading 0's

        f1_data = [R46, R45, R44, R37, R36, R20, R19, R14, R8, R0]  # put freq1 data in array


        index = 2  # start at index 2

        # put data in spreadsheet
        for d in f1_data:
            self.sheet['A'+str(index)] = d
            self.sheet['B'+str(index)] = 0
            self.sheet['C'+str(index)] = 0
            self.sheet['D'+str(index)] = 0
            self.sheet['E'+str(index)] = 0
            index = index + 1
        # frequency 1 again
        # put data in spreadsheet again
        for d in f1_data:
            self.sheet['A'+str(index)] = d
            self.sheet['B'+str(index)] = 0
            self.sheet['C'+str(index)] = 0
            self.sheet['D'+str(index)] = 0
            self.sheet['E'+str(index)] = 0
            index = index + 1

        # add labels to describe excel sheet
       #self.sheet['M1'] = "Script by SH"

        self.sheet['M3'] = "Frequency(MHz)"
        self.sheet['N3'] = freq1
        #self.sheet['M4'] = "Frequency 2 (MHz)"
        #self.sheet['N4'] = freq2
        #self.sheet['M5'] = "Actual Delay (uS)"
        #self.sheet['N5'] = delayBtoA*self.framerate/1000


        self.wb.save(self.filename)  # save file after modification


